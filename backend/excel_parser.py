import io
from typing import Dict, Optional, Any

try:
    import openpyxl
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False

try:
    import pandas as pd
    PANDAS_OK = True
except ImportError:
    PANDAS_OK = False


SHEET_HIST_ENERGY = "historical_energy"
SHEET_HIST_DEMO   = "historical_demo"
SHEET_ELEC_MIX    = "elec_mix"
SHEET_SCENARIOS   = "scenarios"
SHEET_NDC         = "ndc_targets"

REQUIRED_SHEETS = [SHEET_HIST_ENERGY, SHEET_HIST_DEMO, SHEET_ELEC_MIX, SHEET_SCENARIOS, SHEET_NDC]



def parse_kzleap_excel(file_bytes: bytes, filename: str = "") -> Dict:

    if not OPENPYXL_OK and not PANDAS_OK:
        return {"error": "openpyxl and pandas are not installed. Please install at least one of them to parse Excel files."}

    try:
        if OPENPYXL_OK:
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
            sheets_found = wb.sheetnames
        else:
            import pandas as pd
            xf = pd.ExcelFile(io.BytesIO(file_bytes))
            sheets_found = xf.sheet_names
    except Exception as e:
        return {"error": f"Failed to open file: {e}"}

    missing = [s for s in REQUIRED_SHEETS if s not in sheets_found]
    if missing:
        return {
            "error": f"Missing sheets in file: {', '.join(missing)}. "
                     f"Use KZLEAP_DATA_TEMPLATE.xlsx.",
            "sheets_found": sheets_found,
        }

    result: Dict[str, Any] = {"source": "kzleap_excel", "filename": filename}


    try:
        he = _read_sheet_as_dicts(wb if OPENPYXL_OK else None,
                                  file_bytes, SHEET_HIST_ENERGY)
        tpes, elec, co2 = {}, {}, {}
        for row in he:
            y = _int(row.get("Year") or row.get("year"))
            if y is None:
                continue
            v_tpes = _float(row.get("TPES_Mtoe") or row.get("tpes_mtoe"))
            v_elec = _float(row.get("Electricity_TWh") or row.get("electricity_twh"))
            v_co2  = _float(row.get("CO2_Mt") or row.get("co2_mt"))
            if v_tpes is not None: tpes[y] = v_tpes
            if v_elec is not None: elec[y] = v_elec
            if v_co2  is not None: co2[y]  = v_co2

        if not tpes and not elec and not co2:
            return {"error": f"'{SHEET_HIST_ENERGY}' is empty or has invalid format. "
                             "Expected columns: Year, TPES_Mtoe, Electricity_TWh, CO2_Mt"}

        result["historical_tpes"] = tpes
        result["historical_elec"] = elec
        result["historical_co2"]  = co2
    except Exception as e:
        return {"error": f"Error parsing '{SHEET_HIST_ENERGY}': {e}"}

    try:
        hd = _read_sheet_as_dicts(wb if OPENPYXL_OK else None,
                                   file_bytes, SHEET_HIST_DEMO)
        pop, working, urban, gdp = {}, {}, {}, {}
        for row in hd:
            y = _int(row.get("Year") or row.get("year"))
            if y is None:
                continue
            v_pop  = _float(row.get("Population_M") or row.get("population_m"))
            v_work = _float(row.get("WorkingAge_pct") or row.get("workingage_pct"))
            v_urb  = _float(row.get("Urban_pct") or row.get("urban_pct"))
            v_gdp  = _float(row.get("GDP_USD_B") or row.get("gdp_usd_b"))
            if v_pop  is not None: pop[y]     = v_pop
            if v_work is not None: working[y] = v_work
            if v_urb  is not None: urban[y]   = v_urb
            if v_gdp  is not None: gdp[y]     = v_gdp

        result["historical_pop"]         = pop
        result["historical_working_age"] = working
        result["historical_urban"]       = urban
        result["historical_gdp"]         = gdp
    except Exception as e:
        return {"error": f"Error parsing '{SHEET_HIST_DEMO}': {e}"}

    try:
        em = _read_sheet_as_dicts(wb if OPENPYXL_OK else None,
                                   file_bytes, SHEET_ELEC_MIX)
        mix = {}
        for row in em:
            src   = str(row.get("Source") or row.get("source") or "").strip().lower()
            share = _float(row.get("Share_pct") or row.get("share_pct"))
            if src and share is not None:
                mix[src] = share
        if not mix:
            return {"error": f"'{SHEET_ELEC_MIX}' is empty. "
                             "Expected columns: Source (coal/gas/hydro/wind/solar/nuclear), Share_pct"}
        result["elec_mix_base"] = mix
    except Exception as e:
        return {"error": f"Error parsing '{SHEET_ELEC_MIX}': {e}"}

    try:
        sc_rows = _read_sheet_as_dicts(wb if OPENPYXL_OK else None,
                                        file_bytes, SHEET_SCENARIOS)
        scenarios: Dict[str, Dict] = {"BAU": {}, "MT": {}, "DD": {}}
        for row in sc_rows:
            param = str(row.get("Parameter") or row.get("parameter") or "").strip()
            if not param:
                continue
            for key in ["BAU", "MT", "DD"]:
                raw = row.get(key) or row.get(key.lower())
                if raw is not None and str(raw).strip() != "":
                    
                    if param in ("name", "description"):
                        scenarios[key][param] = str(raw).strip()
                    else:
                        v = _float(raw)
                        if v is not None:
                            scenarios[key][param] = v

        # Проверка обязательных полей
        required_params = ["gdp_growth", "renewables_2030", "renewables_2050",
                           "coal_share_2050", "co2_price_2030", "co2_price_2050"]
        for key in ["BAU", "MT", "DD"]:
            missing_p = [p for p in required_params if p not in scenarios[key]]
            if missing_p:
                return {"error": f"Сценарий {key}: отсутствуют параметры {missing_p}"}

  
        for key in ["BAU", "MT", "DD"]:
            scenarios[key].setdefault("nuclear_gw_2035", 0.0)
            scenarios[key].setdefault("ev_share_2050", 0.0)
            scenarios[key].setdefault("energy_intensity_change", -0.015)
            scenarios[key].setdefault("name", key)
            scenarios[key].setdefault("description", f"{key} scenario")

        result["scenarios"] = scenarios
    except Exception as e:
        return {"error": f"Error parsing '{SHEET_SCENARIOS}': {e}"}


    try:
        ndc_rows = _read_sheet_as_dicts(wb if OPENPYXL_OK else None,
                                         file_bytes, SHEET_NDC)
        ndc: Dict[str, Any] = {}
        for row in ndc_rows:
            param = str(row.get("Parameter") or row.get("parameter") or "").strip()
            val   = row.get("Value") or row.get("value")
            if param and val is not None:
                ndc[param] = _float(val) if param != "description" else str(val)

        ndc.setdefault("base_year", 1990)
        ndc.setdefault("base_co2_mt", 290.0)
        ndc.setdefault("ndc_unconditional_pct", -15.0)
        ndc.setdefault("ndc_conditional_pct", -25.0)
        ndc.setdefault("neutrality_year", 2060)

        result["ndc_targets"] = {
            "base_year":              int(ndc["base_year"]),
            "base_co2":               float(ndc["base_co2_mt"]),
            "unconditional_2030_pct": float(ndc["ndc_unconditional_pct"]),
            "conditional_2030_pct":   float(ndc["ndc_conditional_pct"]),
            "neutrality_year":        int(ndc["neutrality_year"]),
        }
    except Exception as e:
        return {"error": f"Error parsing '{SHEET_NDC}': {e}"}

    if result["historical_co2"]:
        last_year = max(result["historical_co2"].keys())
        result["base_year"]  = last_year
        result["base_co2"]   = result["historical_co2"].get(last_year, 242.0)
        result["base_elec"]  = result["historical_elec"].get(last_year, 115.0)
        result["base_tpes"]  = result["historical_tpes"].get(last_year, 85.0)
        result["base_pop"]   = result["historical_pop"].get(last_year, 20.1)
        result["base_urban"] = result["historical_urban"].get(last_year, 58.0)
        result["base_working"] = result["historical_working_age"].get(last_year, 66.5)
        result["base_gdp_pc"]  = result["historical_gdp"].get(last_year, 13.0)

    return result


def _read_sheet_as_dicts(wb, file_bytes: bytes, sheet_name: str):

    if wb is not None:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        header_idx = None
        for i, row in enumerate(rows):
            first_cell = str(row[0]).strip() if row[0] is not None else ""
            if first_cell and not first_cell.startswith("#"):
                header_idx = i
                break
        if header_idx is None:
            return []
        headers = [str(h).strip() if h is not None else "" for h in rows[header_idx]]
        result = []
        for row in rows[header_idx + 1:]:
            if all(v is None for v in row):
                continue
            result.append({headers[i]: row[i] for i in range(min(len(headers), len(row)))})
        return result
    else:
        import pandas as pd
        raw = pd.read_excel(io.BytesIO(file_bytes), sheet_name=sheet_name, header=None)
        header_idx = None
        for i, row in raw.iterrows():
            first = str(row.iloc[0]).strip() if row.iloc[0] is not None else ""
            if first and not first.startswith("#") and first != "nan":
                header_idx = i
                break
        if header_idx is None:
            return []
        df = pd.read_excel(io.BytesIO(file_bytes), sheet_name=sheet_name, header=header_idx)
        df.columns = [str(c).strip() for c in df.columns]
        df = df.dropna(how="all")
        return df.to_dict("records")


def _float(v) -> Optional[float]:
    if v is None:
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return None


def _int(v) -> Optional[int]:
    f = _float(v)
    return int(f) if f is not None else None



def create_template_excel() -> bytes:

    if not OPENPYXL_OK:
        raise RuntimeError("openpyxl не установлен")

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    wb.remove(wb.active)  

    HDR_FILL   = PatternFill("solid", fgColor="1D6B48")
    HDR_FONT   = Font(color="FFFFFF", bold=True)
    NOTE_FILL  = PatternFill("solid", fgColor="E8F5E9")
    NOTE_FONT  = Font(color="1D6B48", italic=True)
    thin       = Side(style="thin", color="CCCCCC")
    border     = Border(left=thin, right=thin, top=thin, bottom=thin)

    def header_row(ws, values, row=1):
        for col, val in enumerate(values, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.fill = HDR_FILL; c.font = HDR_FONT
            c.alignment = Alignment(horizontal="center")
            c.border = border

    def data_row(ws, values, row):
        for col, val in enumerate(values, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.border = border
            if col == 1:
                c.font = Font(bold=True)
            c.alignment = Alignment(horizontal="center")

    def note(ws, text, row, col=1):
        c = ws.cell(row=row, column=col, value=text)
        c.fill = NOTE_FILL; c.font = NOTE_FONT

    def auto_width(ws):
        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=8)
            ws.column_dimensions[get_column_letter(col[0].column)].width = max(max_len + 4, 12)

    ws1 = wb.create_sheet(SHEET_HIST_ENERGY)
    note(ws1, "# Исторические данные по энергетике Казахстана. Заполните реальными значениями.", 1)
    header_row(ws1, ["Year", "TPES_Mtoe", "Electricity_TWh", "CO2_Mt"], row=2)
    hist_e = [
        (1990, 105.0, 87.4,  290.0),
        (1995, 75.0,  67.0,  200.0),
        (2000, 57.0,  51.6,  140.0),
        (2005, 68.0,  67.9,  200.0),
        (2010, 74.0,  82.6,  230.0),
        (2015, 77.0,  90.8,  250.0),
        (2018, 76.0,  105.0, 260.0),
        (2020, 78.0,  107.0, 235.0),
        (2021, 81.0,  109.0, 245.0),
        (2022, 83.0,  112.0, 240.0),
        (2023, 85.0,  115.0, 242.0),
    ]
    for i, row_data in enumerate(hist_e, 3):
        data_row(ws1, row_data, i)
    auto_width(ws1)

    ws2 = wb.create_sheet(SHEET_HIST_DEMO)
    note(ws2, "# Демографические и экономические данные. Population_M — миллионы человек. GDP_USD_B — млрд USD.", 1)
    header_row(ws2, ["Year", "Population_M", "WorkingAge_pct", "Urban_pct", "GDP_USD_B"], row=2)
    hist_d = [
        (1990, 16.5, 61.0, 57.0, 26.0),
        (1995, 15.7, 62.0, 55.5, None),
        (2000, 14.9, 63.5, 55.0, 18.3),
        (2005, 15.2, 65.0, 53.5, 57.1),
        (2010, 16.2, 67.0, 53.5, 148.1),
        (2015, 17.7, 68.5, 53.5, 184.4),
        (2020, 19.0, 67.0, 57.5, 171.1),
        (2021, 19.5, 66.8, 57.7, 197.1),
        (2022, 19.8, 66.6, 57.9, 220.6),
        (2023, 20.1, 66.5, 58.0, 261.4),
    ]
    for i, row_data in enumerate(hist_d, 3):
        data_row(ws2, row_data, i)
    auto_width(ws2)

    ws3 = wb.create_sheet(SHEET_ELEC_MIX)
    note(ws3, "# Структура электрогенерации базового года (последний исторический год). Сумма должна = 100%.", 1)
    header_row(ws3, ["Source", "Share_pct"], row=2)
    mix_data = [
        ("coal",    61.0),
        ("gas",     24.0),
        ("hydro",   10.0),
        ("wind",     3.5),
        ("solar",    1.5),
        ("nuclear",  0.0),
    ]
    for i, row_data in enumerate(mix_data, 3):
        data_row(ws3, row_data, i)
    auto_width(ws3)

    ws4 = wb.create_sheet(SHEET_SCENARIOS)
    note(ws4, "# Параметры сценариев. Каждая строка — параметр, столбцы BAU/MT/DD — значения.", 1)
    header_row(ws4, ["Parameter", "BAU", "MT", "DD"], row=2)
    sc_data = [
        ("name",                    "Business as Usual", "Moderate Transition", "Deep Decarbonization"),
        ("description",             "Current policies continue, NDC targets не достигнуты",
                                    "NDC targets are met, moderate push for renewables and efficiency",
                                    "Carbon neutrality by 2060, aggressive renewables and efficiency"),
        ("gdp_growth",              0.040,  0.042,  0.043),
        ("energy_intensity_change", -0.010, -0.020, -0.030),
        ("coal_share_2050",         0.45,   0.25,   0.05),
        ("renewables_2030",         0.08,   0.15,   0.22),
        ("renewables_2050",         0.15,   0.40,   0.70),
        ("nuclear_gw_2035",         0.0,    1.2,    2.4),
        ("co2_price_2030",          5,      20,     50),
        ("co2_price_2050",          10,     50,     150),
        ("ev_share_2050",           0.10,   0.30,   0.80),
    ]
    for i, row_data in enumerate(sc_data, 3):
        data_row(ws4, row_data, i)
    auto_width(ws4)

    ws5 = wb.create_sheet(SHEET_NDC)
    note(ws5, "# Цели NDC Казахстана. ndc_*_pct — в % к базовому году (отрицательные = снижение).", 1)
    header_row(ws5, ["Parameter", "Value"], row=2)
    ndc_data = [
        ("base_year",              1990),
        ("base_co2_mt",            290.0),
        ("ndc_unconditional_pct", -15.0),
        ("ndc_conditional_pct",   -25.0),
        ("neutrality_year",        2060),
    ]
    for i, row_data in enumerate(ndc_data, 3):
        data_row(ws5, row_data, i)
    auto_width(ws5)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
